import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from typing import List
from models import ProfileRecord

def save_to_excel(records: List[ProfileRecord], file_path: str = "output_candidates.xlsx") -> str:
    if not records:
        return "No records provided to export."

    data = [
        {
            "Full Name": r.full_name,
            "Current Title": r.current_title,
            "Current Company": r.current_company,
            "Location": r.location,
            "Headline": r.headline,
            "Skills": ", ".join(r.skills[:8]),
            "Profile URL": r.profile_url,
            "Scraped At": r.scraped_at
        }
        for r in records
    ]

    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Candidates")
        ws = writer.sheets["Candidates"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(df[col].astype(str).map(len).max(), len(col)) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 45)

        for row_idx in range(2, len(df) + 2):
            url_cell = ws.cell(row=row_idx, column=7)
            raw_url = url_cell.value
            if raw_url and str(raw_url).startswith("http"):
                url_cell.hyperlink = raw_url
                url_cell.font = Font(color="0563C1", underline="single")

    return file_path